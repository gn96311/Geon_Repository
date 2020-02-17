import sys
import re
import requests
from ast import literal_eval
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl

keyword = list(map(str, input().split()))

chrome_driver = 'C:\\Users\\Home\\Downloads\\chromedriver'
driver = webdriver.Chrome(chrome_driver)

def coupang_keyword(keyword):
    time.sleep(2)
    result = []
    keyword_list = f"https://www.coupang.com/np/search/autoComplete?callback=jQuery111105373706370684133_1581842397038&keyword={keyword}&_=1581842397044"
    resp = requests.get(keyword_list)
    try:
        diction = literal_eval(resp.text[43:-2])
        for i in diction:
            try:
                result.append(i.get("keyword"))
            except AttributeError:
                result.append('Nope')
    except:
        return result
    return result

def coupang_list(keyword):
    time.sleep(2)
    next_keyword_list = f"https://www.coupang.com/np/search?component=&q={keyword}&channel=user"
    driver.get(next_keyword_list)
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.ID,'searchOptionForm')))
    src = driver.page_source
    soup = BeautifulSoup(src, "html.parser")
    result_2 = []
    try:
        related_words = soup.find("dl", {"class": "search-related-keyword"}).find_all("dd")
        for i in related_words:
            result_2.append((i.get_text()).split())
            try:
                result_2_new = list(set(result_2[0]))
            except UnboundLocalError:
                result_2_new = result_2
    except AttributeError:
        result_2.append('Nope')
    try:
        return result_2_new
    except UnboundLocalError:
        result_2_new = []
        return result_2_new

def item_scout(keyword):
    try:
        item_scout_url = f'https://itemscout.io/keyword/'
        driver.get(item_scout_url)
        search = driver.find_element_by_class_name("input-keyword")
        search.clear()
        search.send_keys(keyword)
        search.send_keys(Keys.RETURN)
        WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located((By.CLASS_NAME,"related-keywords")))
        src = driver.page_source
        soup = BeautifulSoup(src, "html.parser")
        related_words = soup.select('.related-keywords a')
        result_3 = []
        for j in related_words:
            result_3.append((j.get_text().split())[0])
        return result_3
    except:
        return []


def sum_keyword(keyword):
    return (list(set(coupang_list(keyword) + coupang_keyword(keyword) + item_scout(keyword))))
'''
wb = openpyxl.load_workbook('KEYWORD.xlsx')
sheet1 = wb['Sheet1']

for i in range(len(keyword)):
    key_result = sum_keyword(keyword[i])
    sheet1.cell(row=i+2, column=2).value = keyword[i]
    for j in range(len(key_result)):
        sheet1.cell(row=i+2, column=j+3).value = key_result[j] + ','

driver.close()
wb.save('KEYWORD.xlsx')
wb.close()
'''

name = sum_keyword(keyword[0])
naming = ''
count = 0
for j in name:
    naming += (str(j) + ',')
    count += 1
driver.close()

print(naming)
print(count)