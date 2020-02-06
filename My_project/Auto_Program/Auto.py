from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import os
import openpyxl
import sys

wb = openpyxl.load_workbook('C:\\Users\\Home\\VScode_github\\Geon_Repository\\My_project\\Cost_program\\Listing_excel.xlsx')
sheet1 = wb['Sheet1']

executable_path = "C:\\Users\\Home\\Downloads\\chromedriver"
os.environ["webdriver.chrome.driver"] = executable_path

chrome_options = webdriver.ChromeOptions()

chrome_options.add_argument('load-extension=C:\\Users\\Home\\Downloads\\shopngo')

driver = webdriver.Chrome(executable_path=executable_path, chrome_options=chrome_options)

driver.get('http://ishopngo.co.kr/')

shopngo_id = driver.find_element_by_name('email')
shopngo_password = driver.find_element_by_name('password')

shopngo_id.clear()
shopngo_password.clear()


shopngo_id.send_keys('tbfcompany@naver.com')
shopngo_password.send_keys('qawsedrf!!')

time.sleep(1)
shopngo_password.submit()

enter_number = int(sys.stdin.readline())

def uploading(number):
    address = driver.find_element_by_id('targeturl')
    address.clear()
    excel_url = str(sheet1.cell(row=number, column=3).value)
    excel_product_name = str(sheet1.cell(row=number, column=9).value)
    excel_product_price = str(sheet1.cell(row=number, column=24).value)
    address.send_keys(excel_url)
    scrap = driver.find_element_by_xpath("//span[@class='btn btn-primary btn-mg']")
    scrap.click()
    enter_pass = str(sys.stdin.readline())
    if enter_pass == 'ok':
        pass

    brand = driver.find_elements_by_xpath(".//[@id='brand']")
    brand.send_keys("Costoff")
    
    product_name = driver.find_elements_by_id("title_ko")
    product_name.clear()
    product_name.send_keys(excel_product_name)

    product_price = driver.find_elements_by_id("price_ko")
    product_price.clear()
    product_price.send_keys(excel_product_price)

    product_price_coupang = driver.find_elements_by_id("price_ko_coupang")
    product_price_coupang.clear()
    product_price.send_keys(excel_product_price)

uploading(enter_number)